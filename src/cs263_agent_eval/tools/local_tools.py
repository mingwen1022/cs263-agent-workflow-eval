"""Local tools for generated benchmark sources."""

from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from langchain.tools import tool


def _source_path(source_store: dict[str, Any], source_id: str) -> Path | None:
    value = source_store.get(source_id)
    if value is None:
        return None
    return Path(str(value))


def _source_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix == ".docx":
        return "docx"
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    if suffix in {".txt", ".md"}:
        return "text"
    return suffix.lstrip(".") or "unknown"


def _missing_source(source_id: str, source_store: dict[str, Any]) -> str:
    return f"Source '{source_id}' not found. Available sources: {', '.join(sorted(source_store))}"


def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _read_csv_file(path: Path) -> str:
    rows = list(csv.reader(path.read_text(encoding="utf-8").splitlines()))
    return "\n".join(",".join(row) for row in rows)


def _read_docx_file(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        if texts:
            paragraphs.append("".join(texts))
    return "\n".join(paragraphs)


def _read_pdf_file(path: Path) -> str:
    data = path.read_bytes().decode("latin-1", errors="ignore")
    matches = re.findall(r"\((.*?)(?<!\\)\)\s*Tj", data, flags=re.DOTALL)
    lines = []
    for match in matches:
        line = match.replace("\\(", "(").replace("\\)", ")").replace("\\\\", "\\")
        lines.append(line)
    return "\n".join(lines)


def _read_xlsx_file(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            rows.append(",".join(values))
        sheets.append(f"# sheet: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sheets)


def _read_any_file(path: Path) -> str:
    file_type = _source_type(path)
    if file_type == "pdf":
        return _read_pdf_file(path)
    if file_type == "docx":
        return _read_docx_file(path)
    if file_type == "csv":
        return _read_csv_file(path)
    if file_type == "xlsx":
        return _read_xlsx_file(path)
    return _read_text_file(path)


def _read_typed(source_store: dict[str, Any], source_id: str, expected_type: str) -> str:
    path = _source_path(source_store, source_id)
    if path is None:
        return _missing_source(source_id, source_store)
    actual_type = _source_type(path)
    if actual_type != expected_type:
        return f"Source '{source_id}' is type '{actual_type}', not '{expected_type}'."
    return _read_any_file(path)


def make_source_tools(source_store: dict[str, Any]) -> list[Any]:
    @tool
    def read_source(source_id: str) -> str:
        """Read any generated source by ID, regardless of file type."""
        path = _source_path(source_store, source_id)
        if path is None:
            return _missing_source(source_id, source_store)
        return _read_any_file(path)

    @tool
    def list_sources() -> str:
        """List source IDs and file types available for this task."""
        entries = []
        for source_id in sorted(source_store):
            path = _source_path(source_store, source_id)
            entries.append(f"{source_id} ({_source_type(path) if path else 'unknown'})")
        return "\n".join(entries)

    @tool
    def read_text_source(source_id: str) -> str:
        """Read a text or markdown source by ID."""
        path = _source_path(source_store, source_id)
        if path is None:
            return _missing_source(source_id, source_store)
        actual_type = _source_type(path)
        if actual_type not in {"text"}:
            return f"Source '{source_id}' is type '{actual_type}', not text."
        return _read_text_file(path)

    @tool
    def read_csv(source_id: str) -> str:
        """Read a CSV spreadsheet-like source by ID."""
        return _read_typed(source_store, source_id, "csv")

    @tool
    def read_pdf(source_id: str) -> str:
        """Read a PDF source by ID and return extracted text."""
        return _read_typed(source_store, source_id, "pdf")

    @tool
    def read_docx(source_id: str) -> str:
        """Read a Word DOCX source by ID and return extracted text."""
        return _read_typed(source_store, source_id, "docx")

    @tool
    def read_xlsx(source_id: str) -> str:
        """Read an Excel XLSX source by ID and return sheet rows as CSV-like text."""
        return _read_typed(source_store, source_id, "xlsx")

    @tool
    def run_calculation(expression: str) -> str:
        """Evaluate a simple arithmetic expression."""
        safe_globals = {"__builtins__": {}, "round": round, "min": min, "max": max, "abs": abs}
        try:
            return str(eval(expression, safe_globals))  # noqa: S307
        except Exception as exc:  # pragma: no cover
            return f"Calculation error: {exc}"

    @tool
    def lookup_vendor_status(vendor_name: str) -> str:
        """Look up a vendor's administrative onboarding status."""
        return (
            f"{vendor_name}: vendor onboarding status service only reports tax-form "
            "and insurance paperwork. It does not contain pricing, SLA, incident, "
            "expense, or reimbursement facts."
        )

    @tool
    def search_employee_directory(name: str) -> str:
        """Search an employee directory by name."""
        return (
            f"Directory result for {name}: this tool returns contact routing only. "
            "It does not contain approvals, policy exceptions, incident timelines, "
            "or contract terms."
        )

    @tool
    def check_public_calendar(date: str) -> str:
        """Check public company calendar events for a date."""
        return (
            f"Calendar lookup for {date}: public event listings only. It does not "
            "contain source evidence for the benchmark task."
        )

    @tool
    def estimate_shipping_cost(weight_lbs: float) -> str:
        """Estimate package shipping costs."""
        return (
            f"Estimated shipping cost for {weight_lbs} lb is unavailable for this "
            "benchmark. Shipping rules are unrelated to the current task."
        )

    return [
        list_sources,
        read_source,
        read_text_source,
        read_csv,
        read_pdf,
        read_docx,
        read_xlsx,
        run_calculation,
        lookup_vendor_status,
        search_employee_directory,
        check_public_calendar,
        estimate_shipping_cost,
    ]
